#!/usr/bin/env python3
"""
신규 매장 후보 네이버 검색 검증 + 좌표 조회.
온라인 스토어 / 서울 외 매장 필터링.
결과: scripts/output/신규매장_검증.xlsx
"""
import os, re, time
from pathlib import Path
from dotenv import load_dotenv
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv(Path(__file__).parent / ".env")
SEARCH_ID     = os.environ["NAVER_SEARCH_CLIENT_ID"]
SEARCH_SECRET = os.environ["NAVER_SEARCH_SECRET"]
GEO_ID        = os.environ["NAVER_GEO_CLIENT_ID"]
GEO_SECRET    = os.environ["NAVER_GEO_SECRET"]

OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# 검증할 매장 후보 (이름, 예상지역 or "")
CANDIDATES = [
    ("키토빵앗간",       "마포구"),
    ("로어앤모어",       "서울"),
    ("제로카페 홍대점",  "마포구"),
    ("인터랙트",         "강남구"),
    ("필로베이커리",     "서울"),
    ("sans카페",         "종로구"),
    ("슈퍼말차더카페",   "서울"),
    ("히포디저트",       "서울"),
    ("희한한제과점",     "서울"),
    ("nuldam",           "서울"),
    ("도야팡",           "서울"),
    ("틴베이크",         "서울"),
    ("스윗트윅",         "서울"),
    ("망넛이네",         "서울"),
    ("리솜브레드",       "서울"),
    ("브레드믿유",       "서울"),
    ("윤달",             "서울"),
    ("스플래쉬베이커리", "서울"),
    ("콩지니빵",         "서울"),
    ("미니데일리베이커리","서울"),
    ("잇포레스트",       "서울"),
    ("키에리",           "서울"),
    ("라므아르",         "서울"),
    ("미니마이즈",       "서울"),
    ("어나더쏭",         "서울"),
    ("베이크샵 피봇",    "서울"),
    ("진켈란젤로",       "서울"),
    ("왓더훅",           "서울"),
    ("마이굿밀",         "서울"),
    ("무슈콘",           "서울"),
    ("비건 빵빵이네",    "서울"),
    ("청담 차선책",      "강남구"),
    ("유스트레스",       "서울"),
    ("행복한단백질",     "성동구"),
    ("당가라 과자점",    "서울"),
    ("이로운제과",       "서울"),
    ("스윗프로비전",     "관악구"),
    ("노모어슈가",       "서울"),
    ("휘게라이프",       "서울"),
    ("매화로와",         "서울"),
    ("멜로우스트릿",     "서울"),
]

def strip_html(t):
    return re.sub(r"<[^>]+>", "", t or "")

def naver_search(name, district):
    q = f"{name} {district}".strip()
    r = requests.get(
        "https://openapi.naver.com/v1/search/local.json",
        params={"query": q, "display": 5},
        headers={"X-Naver-Client-Id": SEARCH_ID, "X-Naver-Client-Secret": SEARCH_SECRET},
        timeout=6,
    )
    return r.json().get("items", [])

def naver_geocode(address):
    r = requests.get(
        "https://maps.apigw.ntruss.com/map-geocode/v2/geocode",
        params={"query": address},
        headers={"X-NCP-APIGW-API-KEY-ID": GEO_ID, "X-NCP-APIGW-API-KEY": GEO_SECRET},
        timeout=6,
    )
    addrs = r.json().get("addresses", [])
    if addrs:
        a = addrs[0]
        return float(a["y"]), float(a["x"]), a.get("roadAddress", address)
    return None, None, address

def is_online_only(item):
    """도로명주소가 없으면 온라인 전용으로 간주"""
    return not item.get("roadAddress")

def is_seoul(item):
    addr = item.get("roadAddress") or item.get("address") or ""
    return "서울" in addr

results = []

for name, district in CANDIDATES:
    print(f"[검색] {name} {district}", end=" ... ", flush=True)
    items = []
    try:
        items = naver_search(name, district)
    except Exception as e:
        print(f"(검색오류:{e})", end=" ")

    # 첫 번째 서울 매장 선택 (온라인 제외)
    matched = None
    for item in items:
        if not is_online_only(item) and is_seoul(item):
            matched = item
            break

    if not matched:
        # 서울 외면 첫 번째라도 기록 (manual 검토용)
        matched = items[0] if items else None

    if not matched:
        print("❌ 미등록")
        results.append({
            "name": name, "hint": district,
            "verified_name": "", "road_address": "", "category": "",
            "phone": "", "lat": None, "lng": None, "status": "❌ 미등록"
        })
        time.sleep(0.2)
        continue

    road_addr = matched.get("roadAddress") or matched.get("address") or ""
    verified_name = strip_html(matched.get("title", name))
    phone = matched.get("telephone", "")
    category = matched.get("category", "")
    online = is_online_only(matched)
    seoul = is_seoul(matched)

    lat, lng = None, None
    try:
        if road_addr:
            lat, lng, road_addr = naver_geocode(road_addr)
    except Exception as e:
        print(f"(좌표오류:{e})", end=" ")

    if online:
        status = "⚠️ 온라인전용"
    elif not seoul:
        status = "⚠️ 서울외"
    elif lat:
        status = "✅ 확인"
    else:
        status = "⚠️ 좌표없음"

    print(f"{status} | {verified_name} | {road_addr}")
    results.append({
        "name": name, "hint": district,
        "verified_name": verified_name, "road_address": road_addr,
        "category": category, "phone": phone,
        "lat": lat, "lng": lng, "status": status
    })
    time.sleep(0.3)

# ── Excel 출력 ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "신규매장_검증"

HEADERS = ["입력명", "힌트지역", "네이버확인명", "도로명주소", "카테고리", "전화번호", "lat", "lng", "상태"]
WIDTHS  = [20, 8, 26, 44, 20, 16, 14, 14, 12]

thin = Side(style="thin", color="D0D0D0")
bord = Border(left=thin, right=thin, top=thin, bottom=thin)
hfill = PatternFill("solid", fgColor="1A1D21")
hfont = Font(bold=True, color="FFFFFF", size=11)

for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
    c = ws.cell(1, ci, h)
    c.font, c.fill = hfont, hfill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bord
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 26
ws.freeze_panes = "A2"

STATUS_COLOR = {
    "✅ 확인":   "D9EFD9",
    "⚠️ 온라인전용": "FFF3CD",
    "⚠️ 서울외":  "FDEBD0",
    "⚠️ 좌표없음": "FCF3CF",
    "❌ 미등록":  "FADBD8",
}

for ri, r in enumerate(results, 2):
    row_data = [r["name"], r["hint"], r["verified_name"], r["road_address"],
                r["category"], r["phone"], r["lat"], r["lng"], r["status"]]
    fc = STATUS_COLOR.get(r["status"], "FFFFFF")
    for ci, val in enumerate(row_data, 1):
        c = ws.cell(ri, ci, val)
        c.fill = PatternFill("solid", fgColor=fc)
        c.border = bord
        c.alignment = Alignment(vertical="center")
        if ci in (7, 8) and val is not None:
            c.number_format = "0.0000000"
    ws.row_dimensions[ri].height = 18

xlsx_path = OUTPUT_DIR / "신규매장_검증.xlsx"
wb.save(xlsx_path)

ok  = sum(1 for r in results if r["status"] == "✅ 확인")
warn = sum(1 for r in results if r["status"].startswith("⚠️"))
fail = sum(1 for r in results if r["status"] == "❌ 미등록")
print(f"\n✅ 확인: {ok}  ⚠️ 요검토: {warn}  ❌ 미등록: {fail} / 총 {len(results)}")
print(f"📊 Excel: {xlsx_path}")
