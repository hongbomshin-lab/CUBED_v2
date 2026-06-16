#!/usr/bin/env python3
"""
저당맵 매장 좌표 변환 + INSERT SQL + xlsx 생성.

Step 1: 네이버 로컬 검색 → 도로명주소 검증 + 매장명 확인
Step 2: 네이버 Geocoding → 정밀 좌표 (x=lng, y=lat)
Step 3: INSERT SQL 생성
Step 4: xlsx 생성

실행: python3 scripts/geocode_stores.py
결과: scripts/output/ 폴더에 .sql + .xlsx 저장
"""
import os, sys, json, time, re
from pathlib import Path
from dotenv import load_dotenv
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv(Path(__file__).parent / ".env")

SEARCH_ID     = os.environ["NAVER_SEARCH_CLIENT_ID"]
SEARCH_SECRET = os.environ["NAVER_SEARCH_SECRET"]
GEO_ID        = os.environ["NAVER_GEO_CLIENT_ID"]
GEO_SECRET    = os.environ["NAVER_GEO_SECRET"]

OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# ── 원본 데이터 ─────────────────────────────────────────────────────────
# (no, name, store_type, address_hint, district, phone, instagram_handle, description)
RAW = [
    (1,  "SANS 익선",                 "cafe",       "서울 종로구 수표로28길 30",                   "종로구", "070-7779-1806", "sans.foundry",          "제로슈가 호밀·티라미수, 제로카페인 대체커피"),
    (2,  "차이채",                    "cafe",       "서울 강남구 언주로155길 14 1층",              "강남구", "",              "",                       "제로슈가·글루텐프리·식이섬유 케이크"),
    (3,  "인터랙트 with 글루텐프리",  "cafe",       "서울 강남구 강남대로118길 16 1층",            "강남구", "0507-1373-7290","cafe_interact",          "무설탕·박가루0%·식이섬유, 알룰로스·에리스리톨"),
    (4,  "저스트레스",                "cafe",       "서울 성동구 무학자7가길 9",                   "성동구", "070-7543-4510", "",                       "제로슈거 호밀·크림브륄레, 대체감미료"),
    (5,  "프로베 베이커리",            "cafe",       "서울 마포구 학습로2길 17-22 1층",             "마포구", "0507-1381-7786","",                       "제로슈거·무설탕 카나리·오코, 스테비아"),
    (6,  "제로카페 하나음",            "cafe",       "서울 마포구 이상수로27길 17 지하1층",          "마포구", "0507-1397-7422","",                       "제로슈거 허숭던·이보 도넛"),
    (7,  "행복한 단백질",              "cafe",       "서울 성동구 마조로 58",                       "성동구", "02-2298-4722",  "",                       "저당 고단백 사료·프로틴 도너"),
    (8,  "키토빵식가",                "cafe",       "서울 마포구 상암컵로19길 36",                 "마포구", "",              "ketobbang",              "무설탕·식이섬유·키토·글루텐프리 베이커리"),
    (9,  "무화당",                    "delivery",   "서울 은평구 팬령로77길 28",                   "은평구", "",              "muhwadang",              "저당·식이섬유·무설탕 온라인 베이커리"),
    (10, "스위트프로비전",            "cafe",       "서울 광진구 화이1길 5",                       "광진구", "",              "",                       "키토·글루텐프리 베이커리"),
    (11, "써니브레드",                "cafe",       "서울 성동구 서울숲2길 24-8",                  "성동구", "",              "sunnybread_official",    "식이섬유·키토·글루텐프리 베이커리"),
    (12, "1994 Bakeshop",             "cafe",       "서울 마포구 마포대로 36",                     "마포구", "",              "1994_bakeshop",          "Low Sugar & Gluten Free 베이커리"),
    (13, "제로랩 용산역점",            "zero_store", "서울 용산구 한강대로 69",                    "용산구", "",              "zerolab.official",       "저당·저칼로리·제로 전문 무인스토어"),
    (14, "제로랩 수유역점",            "zero_store", "서울 강북구 도봉로 342 1층",                  "강북구", "",              "zerolab.official",       "저당·저칼로리·제로 전문 무인스토어"),
    (15, "제로랩 남성역점",            "zero_store", "서울 동작구 삼성로 280 1층",                  "동작구", "",              "zerolab.official",       "저당·저칼로리·제로 전문 무인스토어"),
    (16, "제로랩 상계역점",            "zero_store", "서울 노원구 한글비석로 396",                  "노원구", "",              "zerolab.official",       "저당·저칼로리·제로 전문 무인스토어"),
    (17, "제로랩 강서점",              "zero_store", "서울 강서구 강서로 344 지하1층",              "강서구", "",              "zerolab.official",       "저당·저칼로리·제로 전문 무인스토어"),
    (18, "제로랩 어린이대공원역점",    "zero_store", "서울 광진구 능동로 181 1층",                  "광진구", "",              "zerolab.official",       "저당·저칼로리·제로 전문 무인스토어"),
    (19, "제로랩 롯데월드몰점",        "zero_store", "서울 송파구 올림픽로 300",                   "송파구", "",              "zerolab.official",       "저당·저칼로리·제로 전문 무인스토어"),
    (20, "제로랩 구로구청점",          "zero_store", "서울 구로구 가마산로 271",                   "구로구", "",              "zerolab.official",       "저당·저칼로리·제로 전문 무인스토어"),
    (21, "제로랩 세곡점",              "zero_store", "서울 강남구 헌릉로571길 13",                 "강남구", "",              "zerolab.official",       "저당·저칼로리·제로 전문 무인스토어"),
    (22, "제로랩 광화문점",            "zero_store", "서울 종로구 세종대로 175",                   "종로구", "",              "zerolab.official",       "저당·저칼로리·제로 전문 무인스토어"),
    (23, "제로랩 신림점",              "zero_store", "서울 금천구 남부순환로 1432",                "금천구", "",              "zerolab.official",       "저당·저칼로리·제로 전문 무인스토어"),
    (24, "제로스토어 신림점",          "zero_store", "서울 관악구 남부순환로 1605",                "관악구", "",              "zerostore.kr",           "저당·저칼로리·제로 전문 무인스토어"),
    # ── 신규 확인 매장 (verify_new_stores.py 기준) ──────────────────────────
    (25, "키토빵앗간",               "cafe",       "서울 마포구 월드컵로19길 36",                 "마포구", "",              "ketobbang",              "키토·글루텐프리·무설탕 베이커리"),
    (26, "제로카페 홍대점",           "cafe",       "서울 마포구 와우산로27길 17",                 "마포구", "",              "",                       "제로슈거 카페·베이커리"),
    (27, "필로베이커리",              "cafe",       "서울 마포구 홍익로2길 27-22",                 "마포구", "",              "",                       "저당·무설탕 베이커리"),
    (28, "히포디저트",               "cafe",       "서울 광진구 아차산로59길 38",                 "광진구", "",              "",                       "저당 디저트"),
    (29, "널담은공간 경복궁점",       "cafe",       "서울 종로구 삼청로 24",                       "종로구", "",              "nuldam",                 "저당·비건 카페"),
    (30, "스윗트윅",                 "cafe",       "서울 마포구 성미산로31길 13",                 "마포구", "",              "",                       "설탕·밀·쌀가루 무첨가 베이커리"),
    (31, "키에리",                   "cafe",       "서울 용산구 이태원로26길 16-8",               "용산구", "",              "",                       "저당·글루텐프리 베이커리"),
    (32, "라므아르",                 "cafe",       "서울 마포구 월드컵로11길 46-14",              "마포구", "",              "",                       "저당 베이커리"),
    (33, "어나더쏭 중곡점",           "cafe",       "서울 광진구 능동로48길 57",                   "광진구", "",              "",                       "저당 베이커리·카페"),
    (34, "진켈란젤로",               "cafe",       "서울 마포구 독막로8길 10",                    "마포구", "",              "",                       "저당 치즈케이크 전문"),
    (35, "왓더훅",                   "cafe",       "서울 강북구 오패산로30길 55",                 "강북구", "",              "",                       "저당 베이커리"),
    (36, "차선책 압구정점",           "cafe",       "서울 강남구 선릉로155길 14",                  "강남구", "",              "",                       "저당 케이크·디저트"),
    (37, "당가라과자점",              "cafe",       "서울 강북구 오현로6길 34",                    "강북구", "",              "",                       "저당·무설탕 과자·베이커리"),
    (38, "이로운제과",               "cafe",       "서울 광진구 뚝섬로32길 29",                   "광진구", "",              "",                       "저당·건강 제과"),
    (39, "스윗프로비전 낙성대점",     "cafe",       "서울 관악구 행운1길 5",                       "관악구", "",              "",                       "키토·글루텐프리 베이커리"),
    (40, "휘게라이프",               "cafe",       "서울 관악구 남현길 37",                       "관악구", "",              "",                       "저당·비건 베이커리·카페"),
]

def strip_html(t):
    return re.sub(r"<[^>]+>", "", t or "")

def naver_local_search(query: str) -> dict | None:
    r = requests.get(
        "https://openapi.naver.com/v1/search/local.json",
        params={"query": query, "display": 3},
        headers={"X-Naver-Client-Id": SEARCH_ID, "X-Naver-Client-Secret": SEARCH_SECRET},
        timeout=6,
    )
    items = r.json().get("items", [])
    return items[0] if items else None

def naver_geocode(address: str) -> tuple[float | None, float | None, str]:
    r = requests.get(
        "https://maps.apigw.ntruss.com/map-geocode/v2/geocode",
        params={"query": address},
        headers={"X-NCP-APIGW-API-KEY-ID": GEO_ID, "X-NCP-APIGW-API-KEY": GEO_SECRET},
        timeout=6,
    )
    addrs = r.json().get("addresses", [])
    if addrs:
        a = addrs[0]
        return float(a["y"]), float(a["x"]), a.get("roadAddress", address)
    return None, None, address

# ── 처리 루프 ─────────────────────────────────────────────────────────────
results, failed = [], []

for no, name, store_type, addr_hint, district, phone, insta, desc in RAW:
    print(f"[{no:02d}] {name}", end=" ... ", flush=True)

    road_addr = addr_hint
    verified_name = name
    tel = phone

    # Step 1: 로컬 검색으로 정확한 도로명주소·매장명·전화번호 확인
    try:
        item = naver_local_search(f"{name} {district}")
        if item:
            road_addr    = item.get("roadAddress") or item.get("address") or addr_hint
            verified_name = strip_html(item.get("title", name))
            tel          = item.get("telephone", "") or phone
    except Exception as e:
        print(f"(검색오류:{e})", end=" ")

    # Step 2: Geocoding으로 정밀 좌표
    lat, lng = None, None
    try:
        lat, lng, road_addr = naver_geocode(road_addr)
    except Exception as e:
        print(f"(좌표오류:{e})", end=" ")

    status = "✅" if lat else "❌"
    print(f"{status} {verified_name} | lat={lat}, lng={lng}")

    if not lat:
        failed.append((no, name, road_addr))

    results.append({
        "no": no,
        "name": verified_name,
        "store_type": store_type,
        "address": road_addr,
        "district": district,
        "lat": lat,
        "lng": lng,
        "phone": tel or "",
        "instagram_url": f"https://instagram.com/{insta}" if insta else "",
        "description": desc,
        "verified": status,
    })
    time.sleep(0.25)  # rate limit

print(f"\n✅ 좌표 획득: {sum(1 for r in results if r['lat'])}/{len(results)}")
if failed:
    print("❌ 실패:", [(f[0], f[1]) for f in failed])

# ── INSERT SQL ────────────────────────────────────────────────────────────
def sv(v):
    if v is None:
        return "NULL"
    if isinstance(v, float):
        return str(v)
    return "'" + str(v).replace("'", "''") + "'"

sql_rows = ",\n".join(
    f"  ({sv(r['name'])},{sv(r['address'])},{sv(r['lat'])},{sv(r['lng'])},"
    f"{sv(r['phone'])},{sv(r['instagram_url'])},{sv(r['store_type'])},{sv(r['district'])},true)"
    for r in results
)
sql = "\n".join([
    "-- 저당맵 stores INSERT SQL",
    "-- 생성: geocode_stores.py (네이버 검색+Geocoding 기준)",
    "-- ❌ 표시 행(lat=NULL)은 수동 보완 필요",
    "",
    "INSERT INTO public.stores",
    "  (name, address, lat, lng, phone, instagram_url, store_type, district, is_active)",
    "VALUES",
    sql_rows + ";",
])
sql_path = OUTPUT_DIR / "stores_insert.sql"
sql_path.write_text(sql, encoding="utf-8")
print(f"📄 SQL: {sql_path}")

# ── Excel ─────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "DB입력_저당매장"

HEADERS = ["No", "매장명(네이버기준)", "store_type", "도로명주소", "구",
           "lat", "lng", "전화번호", "인스타그램URL", "설명", "네이버검증"]
WIDTHS  = [4, 28, 13, 45, 8, 14, 14, 16, 38, 48, 8]

thin  = Side(style="thin", color="D0D0D0")
bord  = Border(left=thin, right=thin, top=thin, bottom=thin)
hfill = PatternFill("solid", fgColor="1A1D21")
hfont = Font(bold=True, color="FFFFFF", size=11)

for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
    c = ws.cell(1, ci, h)
    c.font, c.fill = hfont, hfill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bord
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 26
ws.freeze_panes = "A2"

TYPE_COLOR = {
    "cafe":       "D9EFD9",
    "restaurant": "FDEBD0",
    "zero_store": "D6EAF8",
    "delivery":   "E8DAEF",
}

for ri, r in enumerate(results, 2):
    row_data = [r["no"], r["name"], r["store_type"], r["address"], r["district"],
                r["lat"], r["lng"], r["phone"], r["instagram_url"], r["description"], r["verified"]]
    fc = TYPE_COLOR.get(r["store_type"], "FFFFFF")
    for ci, val in enumerate(row_data, 1):
        c = ws.cell(ri, ci, val)
        c.fill   = PatternFill("solid", fgColor=fc)
        c.border = bord
        c.alignment = Alignment(vertical="center", wrap_text=(ci >= 9))
        if ci in (6, 7) and val is not None:
            c.number_format = "0.0000000"
    if r["verified"] == "❌":
        for ci in range(1, 12):
            ws.cell(ri, ci).font = Font(color="C0392B", bold=True)
    ws.row_dimensions[ri].height = 18

# 범례
ws2 = wb.create_sheet("범례")
for ci, h in enumerate(["store_type", "색상", "설명"], 1):
    ws2.cell(1, ci, h).font = Font(bold=True)
for ri, (t, label) in enumerate([
    ("cafe", "카페·베이커리"),
    ("restaurant", "음식점"),
    ("zero_store", "제로스토어"),
    ("delivery", "배달"),
], 2):
    for ci, v in enumerate([t, TYPE_COLOR[t], label], 1):
        ws2.cell(ri, ci, v).fill = PatternFill("solid", fgColor=TYPE_COLOR[t])

# 실패 시트
if failed:
    ws3 = wb.create_sheet("좌표실패_수동보완")
    for ci, h in enumerate(["No", "매장명", "검색주소"], 1):
        ws3.cell(1, ci, h).font = Font(bold=True)
    for ri, (no, nm, addr) in enumerate(failed, 2):
        ws3.cell(ri, 1, no); ws3.cell(ri, 2, nm); ws3.cell(ri, 3, addr)

xlsx_path = OUTPUT_DIR / "저당맵_DB입력_검증.xlsx"
wb.save(xlsx_path)
print(f"📊 Excel: {xlsx_path}")

# ── 미확인 매장 CSV ────────────────────────────────────────────────────────
import csv

UNVERIFIED = [
    # 기존 RAW에서 좌표 실패
    ("차이채",           "강남구", "서울 강남구 언주로155길 14 1층",         "❌ 네이버 미등록"),
    ("저스트레스",       "성동구", "서울 성동구 무학자7가길 9",               "❌ 네이버 미등록"),
    ("프로베 베이커리",  "마포구", "서울 마포구 학습로2길 17-22 1층",        "❌ 네이버 미등록"),
    ("제로카페 하나음",  "마포구", "서울 마포구 이상수로27길 17 지하1층",     "❌ 네이버 미등록"),
    ("무화당",           "은평구", "서울 은평구 팬령로77길 28",               "⚠️ 서울 지점 미등록(일산점만 확인)"),
    ("제로랩 남성역점",  "동작구", "서울 동작구 삼성로 280 1층",             "❌ 네이버 미등록"),
    # 신규 후보에서 미등록
    ("로어앤모어",       "",      "",                                         "❌ 네이버 미등록"),
    ("슈퍼말차더카페",   "",      "",                                         "❌ 네이버 미등록"),
    ("희한한제과점",     "",      "",                                         "❌ 네이버 미등록"),
    ("도야팡",           "",      "",                                         "❌ 네이버 미등록"),
    ("틴베이크",         "",      "",                                         "❌ 네이버 미등록"),
    ("리솜브레드",       "",      "",                                         "❌ 네이버 미등록"),
    ("브레드믿유",       "",      "",                                         "❌ 네이버 미등록"),
    ("스플래쉬베이커리", "",      "",                                         "❌ 네이버 미등록"),
    ("콩지니빵",         "",      "",                                         "❌ 네이버 미등록"),
    ("미니데일리베이커리","",     "",                                         "❌ 네이버 미등록"),
    ("잇포레스트",       "",      "",                                         "❌ 네이버 미등록"),
    ("베이크샵 피봇",    "",      "",                                         "❌ 네이버 미등록"),
    ("마이굿밀",         "",      "",                                         "❌ 네이버 미등록"),
    ("무슈콘",           "",      "",                                         "❌ 네이버 미등록"),
    ("비건 빵빵이네",    "",      "",                                         "❌ 네이버 미등록"),
    ("유스트레스",       "",      "",                                         "❌ 네이버 미등록"),
    ("행복한단백질",     "성동구","서울 성동구 마조로 58",                    "❌ 네이버 미등록"),
    ("노모어슈가",       "",      "",                                         "❌ 네이버 미등록"),
    ("매화로와",         "",      "",                                         "❌ 네이버 미등록"),
    ("멜로우스트릿",     "",      "",                                         "❌ 네이버 미등록"),
    # 오매칭 (검색결과가 다른 업종)
    ("망넛이네",         "서대문구","",                                       "⚠️ 오매칭(수버킷=기업) — 주소 직접 입력 필요"),
    ("윤달",             "",      "",                                         "⚠️ 오매칭(윤달오리전문점=식당) — 재확인 필요"),
    ("미니마이즈",       "",      "",                                         "⚠️ 오매칭(워크미니마이즈=사무공간) — 재확인 필요"),
]

csv_path = OUTPUT_DIR / "미확인매장_수동보완필요.csv"
with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["매장명", "예상구", "알고있는주소", "사유"])
    w.writerows(UNVERIFIED)
print(f"📋 미확인 CSV: {csv_path}")
